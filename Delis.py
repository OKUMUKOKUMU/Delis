# app.py - Streamlit Invoice Extraction App

import streamlit as st
import pandas as pd
import numpy as np
import re
import io
import warnings
import base64
from datetime import datetime
import time
import sys

# Try to import openpyxl, show instructions if not installed
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    st.error("⚠️ **openpyxl is not installed!** Please install it using: `pip install openpyxl`")

warnings.filterwarnings('ignore')

# Set page configuration
st.set_page_config(
    page_title="Invoice Data Extractor",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for better styling
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        color: #1E3A8A;
        text-align: center;
        margin-bottom: 1rem;
    }
    .sub-header {
        font-size: 1.5rem;
        color: #374151;
        margin-top: 1.5rem;
        margin-bottom: 1rem;
    }
    .success-box {
        background-color: #D1FAE5;
        padding: 1rem;
        border-radius: 0.5rem;
        border-left: 4px solid #10B981;
        margin: 1rem 0;
    }
    .info-box {
        background-color: #DBEAFE;
        padding: 1rem;
        border-radius: 0.5rem;
        border-left: 4px solid #3B82F6;
        margin: 1rem 0;
    }
    .warning-box {
        background-color: #FEF3C7;
        padding: 1rem;
        border-radius: 0.5rem;
        border-left: 4px solid #F59E0B;
        margin: 1rem 0;
    }
    .error-box {
        background-color: #FEE2E2;
        padding: 1rem;
        border-radius: 0.5rem;
        border-left: 4px solid #EF4444;
        margin: 1rem 0;
    }
    .metric-card {
        background-color: #F8FAFC;
        padding: 1rem;
        border-radius: 0.5rem;
        border: 1px solid #E2E8F0;
        text-align: center;
    }
    .stProgress > div > div > div > div {
        background-color: #3B82F6;
    }
    .install-instructions {
        background-color: #F3F4F6;
        padding: 1.5rem;
        border-radius: 0.5rem;
        border: 2px dashed #6B7280;
        margin: 1rem 0;
        font-family: monospace;
    }
</style>
""", unsafe_allow_html=True)

def check_dependencies():
    """Check if all required packages are installed"""
    missing_packages = []
    
    try:
        import openpyxl
    except ImportError:
        missing_packages.append("openpyxl")
    
    try:
        import pandas
    except ImportError:
        missing_packages.append("pandas")
    
    try:
        import numpy
    except ImportError:
        missing_packages.append("numpy")
    
    return missing_packages

# Check dependencies first
if not OPENPYXL_AVAILABLE:
    st.markdown('<div class="error-box">', unsafe_allow_html=True)
    st.markdown("### ⚠️ Missing Dependencies")
    st.markdown("""
    The following required packages are not installed:
    
    **Required packages:**
    - `openpyxl` (for reading Excel files)
    - `pandas` (for data manipulation)
    - `numpy` (for numerical operations)
    
    **Installation instructions:**
    """)
    
    st.markdown('<div class="install-instructions">', unsafe_allow_html=True)
    st.code("pip install openpyxl pandas numpy", language="bash")
    st.markdown('</div>', unsafe_allow_html=True)
    
    st.markdown("""
    **For Streamlit Cloud deployment**, create a `requirements.txt` file with:
    """)
    
    st.markdown('<div class="install-instructions">', unsafe_allow_html=True)
    st.code("""openpyxl==3.1.2
pandas==2.1.0
numpy==1.24.0
streamlit==1.28.0""", language="txt")
    st.markdown('</div>', unsafe_allow_html=True)
    
    st.markdown("""
    **After installing**, refresh this page or restart the app.
    """)
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Stop execution if dependencies are missing
    st.stop()

# Initialize session state
if 'extracted_data' not in st.session_state:
    st.session_state.extracted_data = None
if 'processing_complete' not in st.session_state:
    st.session_state.processing_complete = False
if 'selected_sheets_count' not in st.session_state:
    st.session_state.selected_sheets_count = 0

def read_excel_with_merged_cells(file_content, sheet_name):
    """Read Excel file and handle merged cells properly"""
    wb = load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb[sheet_name]
    
    # Get merged cell ranges
    merged_ranges = ws.merged_cells.ranges
    
    # Create a dictionary to map merged cells
    merged_cells_map = {}
    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        # Get the value from the top-left cell
        top_left_value = ws.cell(row=min_row, column=min_col).value
        
        # Map all cells in this range to the top-left value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_cells_map[(row, col)] = top_left_value
    
    # Read all data into a list of lists
    data = []
    max_row = ws.max_row
    max_col = ws.max_column
    
    for row in range(1, max_row + 1):
        row_data = []
        for col in range(1, max_col + 1):
            # Check if cell is in merged range
            if (row, col) in merged_cells_map:
                cell_value = merged_cells_map[(row, col)]
            else:
                cell_value = ws.cell(row=row, column=col).value
            
            row_data.append(cell_value)
        data.append(row_data)
    
    # Convert to DataFrame
    df = pd.DataFrame(data)
    
    return df

def clean_numeric_value(value):
    """Clean a numeric value by removing non-numeric characters"""
    if pd.isna(value):
        return value
    
    str_value = str(value)
    # Remove currency symbols, commas, spaces
    str_value = re.sub(r'[^\d\.\-]', '', str_value)
    return str_value

def clean_and_convert_item_data(items_df):
    """Clean and convert item data to proper types"""
    # Clean UoM column
    if 'UoM' in items_df.columns:
        items_df['UoM'] = items_df['UoM'].apply(
            lambda x: str(x).strip().title() if pd.notna(x) and str(x).strip() else 'Kilograms'
        )
    
    # Convert numeric columns
    numeric_columns = ['Qty', 'Unit Price Excl. VAT', 'VAT %', 'Line Amount Excl. VAT']
    
    for col in numeric_columns:
        if col in items_df.columns:
            # Clean the values
            items_df[col] = items_df[col].apply(
                lambda x: clean_numeric_value(x) if pd.notna(x) else None
            )
            # Convert to numeric
            items_df[col] = pd.to_numeric(items_df[col], errors='coerce')
    
    # Calculate missing line amounts
    if 'Line Amount Excl. VAT' in items_df.columns and 'Qty' in items_df.columns and 'Unit Price Excl. VAT' in items_df.columns:
        mask = (items_df['Line Amount Excl. VAT'].isna() | (items_df['Line Amount Excl. VAT'] == 0)) & items_df['Qty'].notna() & items_df['Unit Price Excl. VAT'].notna()
        items_df.loc[mask, 'Line Amount Excl. VAT'] = items_df.loc[mask, 'Qty'] * items_df.loc[mask, 'Unit Price Excl. VAT']
    
    # Set default VAT % if missing
    if 'VAT %' in items_df.columns:
        items_df['VAT %'] = items_df['VAT %'].fillna(16.0)
    
    # Remove empty rows
    items_df = items_df[items_df['No.'].astype(str).str.strip() != '']
    items_df = items_df[items_df['Description'].astype(str).str.strip() != '']
    
    return items_df

def extract_financial_totals(df, row_idx, header_positions, result):
    """Extract financial totals from a totals row"""
    for col_name, col_idx in header_positions.items():
        if col_idx < len(df.columns):
            cell_value = df.iloc[row_idx, col_idx]
            if pd.notna(cell_value):
                try:
                    # Clean and convert to number
                    num_str = str(cell_value).replace(',', '').replace(' ', '')
                    num_val = float(num_str)
                    
                    # Determine which total this is based on column
                    if 'Line Amount' in col_name:
                        first_cell = str(df.iloc[row_idx, header_positions.get('No.', 0)]).lower()
                        if 'subtotal' in first_cell:
                            result['subtotal'] = num_val
                        elif 'vat amount' in first_cell:
                            result['vat_amount'] = num_val
                        elif 'total' in first_cell:
                            result['total_amount'] = num_val
                except:
                    pass

def extract_invoice_data_from_sheet(df, file_name, sheet_name, progress_bar=None, status_text=None):
    """Extract invoice data from a single sheet"""
    result = {
        'file_name': file_name,
        'sheet_name': sheet_name,
        'customer_name': '',
        'document_date': '',
        'invoice_number': '',
        'order_number': '',
        'items_df': None,
        'subtotal': 0,
        'vat_amount': 0,
        'total_amount': 0
    }
    
    # 1. Extract Customer/Branch Name
    customer_names = []
    for idx in range(min(20, len(df))):
        row_vals = []
        for col in range(min(15, len(df.columns))):
            cell_val = df.iloc[idx, col] if col < len(df.columns) else ''
            if pd.notna(cell_val) and str(cell_val).strip():
                row_vals.append(str(cell_val).strip())
        
        row_str = ' '.join(row_vals)
        
        # Look for customer names
        if any(keyword in row_str for keyword in ['Chandarana', 'Delis', 'Branch', 'Buffalo', 'Naivasha']):
            name_patterns = [
                r'(Chandarana[^\d\n]{0,50}Branch)',
                r'(Chandarana[^\d\n]{0,50}Mall)',
                r'(Chandarana[^\d\n]{0,50}Naivasha)',
                r'(Chandarana[^\d\n]{0,30})',
            ]
            
            for pattern in name_patterns:
                match = re.search(pattern, row_str, re.IGNORECASE)
                if match:
                    name = match.group(1).strip()
                    if name and name not in customer_names and len(name) > 5:
                        customer_names.append(name)
                        break
    
    if customer_names:
        branch_names = [name for name in customer_names if 'Branch' in name or 'Mall' in name]
        if branch_names:
            result['customer_name'] = branch_names[0]
        else:
            result['customer_name'] = customer_names[-1]
    
    # 2. Extract Document Date
    for idx in range(len(df)):
        for col in range(min(10, len(df.columns))):
            cell_val = str(df.iloc[idx, col])
            
            if cell_val.strip() == 'Document Date':
                # Check columns E, F, G (4, 5, 6 in 0-index)
                target_columns = [4, 5, 6]
                
                for target_col in target_columns:
                    if target_col < len(df.columns):
                        date_cell = df.iloc[idx, target_col]
                        if pd.notna(date_cell):
                            date_str = str(date_cell).strip()
                            
                            date_patterns = [
                                r'(\d{1,2}\.\s+[A-Za-z]+\s+\d{4})',
                                r'(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})',
                                r'(\d{4}[/\-\.]\d{1,2}[/\-\.]\d{1,2})',
                                r'(\d{1,2}-[A-Za-z]{3}-\d{4})',
                                r'(\d{1,2}\s+[A-Za-z]{3,9}\s+\d{4})',
                            ]
                            
                            for pattern in date_patterns:
                                match = re.search(pattern, date_str)
                                if match:
                                    result['document_date'] = match.group(1)
                                    break
    
    # 3. Extract Invoice Number
    invoice_found = False
    for idx in range(len(df)):
        for col in range(min(15, len(df.columns))):
            cell_val = str(df.iloc[idx, col])
            
            if 'Invoice No.' in cell_val and not invoice_found:
                # Check to the RIGHT
                for offset in range(1, 6):
                    check_col = col + offset
                    if check_col < len(df.columns):
                        inv_cell = df.iloc[idx, check_col]
                        if pd.notna(inv_cell):
                            inv_str = str(inv_cell).strip()
                            if (inv_str and len(inv_str) > 3 and 
                                re.match(r'^\d+$', inv_str) and
                                'Invoice' not in inv_str and 
                                'No.' not in inv_str and
                                'CU' not in inv_str):
                                result['invoice_number'] = inv_str
                                invoice_found = True
                                break
    
    # 4. Extract Order Number
    order_found = False
    for idx in range(len(df)):
        for col in range(min(15, len(df.columns))):
            cell_val = str(df.iloc[idx, col])
            
            if 'Order No.' in cell_val and not order_found:
                # Check to the RIGHT
                for offset in range(1, 6):
                    check_col = col + offset
                    if check_col < len(df.columns):
                        order_cell = df.iloc[idx, check_col]
                        if pd.notna(order_cell):
                            order_str = str(order_cell).strip()
                            if order_str and re.match(r'^\d+$', order_str):
                                result['order_number'] = order_str
                                order_found = True
                                break
    
    # 5. Find and extract items table
    items_data = []
    header_positions = {}
    
    # Look for the header row
    for idx in range(len(df)):
        current_header_positions = {}
        
        for col in range(min(20, len(df.columns))):
            cell_val = str(df.iloc[idx, col]).strip().lower() if col < len(df.columns) else ''
            
            header_mappings = {
                'no.': 'No.',
                'description': 'Description',
                'qty': 'Qty',
                'uom': 'UoM',
                'unit price excl. vat': 'Unit Price Excl. VAT',
                'unit price': 'Unit Price Excl. VAT',
                'vat %': 'VAT %',
                'vat%': 'VAT %',
                'line amount excl. vat': 'Line Amount Excl. VAT',
                'line amount': 'Line Amount Excl. VAT'
            }
            
            for keyword, col_name in header_mappings.items():
                if keyword in cell_val:
                    current_header_positions[col_name] = col
                    break
        
        if len(current_header_positions) >= 4:
            header_positions = current_header_positions
            
            # Extract items from rows below
            for item_idx in range(idx + 1, len(df)):
                if 'No.' in header_positions:
                    no_col = header_positions['No.']
                    if no_col < len(df.columns):
                        item_no = df.iloc[item_idx, no_col]
                        
                        if pd.isna(item_no):
                            continue
                        
                        item_no_str = str(item_no).strip()
                        
                        if (re.match(r'^[A-Z]{2,4}-\d{5}', item_no_str) or 
                            re.match(r'^\d+[A-Z]?$', item_no_str)):
                            
                            item_data = {}
                            
                            for col_name, col_idx in header_positions.items():
                                if col_idx < len(df.columns):
                                    cell_value = df.iloc[item_idx, col_idx]
                                    item_data[col_name] = cell_value if pd.notna(cell_value) else ''
                            
                            if 'Description' in item_data:
                                item_data['Description'] = str(item_data['Description']).strip()
                            
                            items_data.append(item_data)
                        
                        elif any(keyword in item_no_str.lower() for keyword in ['subtotal', 'total', 'vat amount']):
                            extract_financial_totals(df, item_idx, header_positions, result)
                            break
            
            break
    
    # 6. Create DataFrame from items
    if items_data:
        items_df = pd.DataFrame(items_data)
        
        required_columns = ['No.', 'Description', 'Qty', 'UoM', 
                           'Unit Price Excl. VAT', 'VAT %', 'Line Amount Excl. VAT']
        
        for col in required_columns:
            if col not in items_df.columns:
                items_df[col] = ''
        
        items_df = clean_and_convert_item_data(items_df)
        result['items_df'] = items_df
        
        if result['subtotal'] == 0 and 'Line Amount Excl. VAT' in items_df.columns:
            result['subtotal'] = items_df['Line Amount Excl. VAT'].sum(skipna=True)
        
        if result['vat_amount'] == 0 and 'Line Amount Excl. VAT' in items_df.columns:
            result['vat_amount'] = items_df['Line Amount Excl. VAT'].sum(skipna=True) * 0.16
        
        if result['total_amount'] == 0:
            result['total_amount'] = result['subtotal'] + result['vat_amount']
    
    return result

def get_excel_download_link(df, filename):
    """Generate a download link for Excel file"""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Extracted Data')
    b64 = base64.b64encode(output.getvalue()).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}">📥 Download Excel File</a>'
    return href

# Sidebar
with st.sidebar:
    st.image("https://cdn-icons-png.flaticon.com/512/3135/3135715.png", width=100)
    st.title("Invoice Extractor")
    st.markdown("---")
    
    st.markdown("### 📋 How to Use:")
    st.markdown("""
    1. **Upload** your Excel invoice file
    2. **Configure** extraction settings
    3. **Process** the file
    4. **Download** the results
    """)
    
    st.markdown("---")
    st.markdown("### ⚙️ Settings")
    
    # Settings
    processing_mode = st.selectbox(
        "Processing Mode",
        ["All Sheets", "First N Sheets", "Last N Sheets", "Custom Range"]
    )
    
    if processing_mode == "First N Sheets":
        n_sheets = st.number_input("Number of sheets", min_value=1, value=10, step=1)
    elif processing_mode == "Last N Sheets":
        n_sheets = st.number_input("Number of sheets", min_value=1, value=10, step=1)
    elif processing_mode == "Custom Range":
        col1, col2 = st.columns(2)
        with col1:
            start_sheet = st.number_input("Start sheet", min_value=1, value=1, step=1)
        with col2:
            end_sheet = st.number_input("End sheet", min_value=1, value=10, step=1)
    
    verbose_mode = st.checkbox("Show detailed extraction logs", value=False)
    show_progress = st.checkbox("Show progress bar", value=True)
    
    st.markdown("---")
    st.markdown("### 📊 Features")
    st.markdown("""
    ✅ Extract from merged cells  
    ✅ Handle multiple date formats  
    ✅ Extract invoice & order numbers  
    ✅ Get all product details  
    ✅ Process multiple sheets  
    ✅ Clean & format data
    """)

# Main content
st.markdown('<h1 class="main-header">📄 Invoice Data Extraction Tool</h1>', unsafe_allow_html=True)
st.markdown("Extract invoice data from Excel files with ease")

# File upload section
st.markdown("### 📤 Upload Your Excel File")
uploaded_file = st.file_uploader(
    "Choose an Excel file",
    type=['xlsx', 'xls'],
    help="Upload Excel files containing invoice data"
)

if uploaded_file:
    st.markdown(f'<div class="success-box"><strong>File uploaded:</strong> {uploaded_file.name}</div>', unsafe_allow_html=True)
    
    try:
        # Get sheet names
        wb = load_workbook(io.BytesIO(uploaded_file.getvalue()), read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        total_sheets = len(sheet_names)
        
        st.markdown(f'<div class="info-box"><strong>Found:</strong> {total_sheets} sheets in the file</div>', unsafe_allow_html=True)
        
        # Determine which sheets to process
        if processing_mode == "All Sheets":
            selected_sheets = sheet_names
            st.session_state.selected_sheets_count = total_sheets
        elif processing_mode == "First N Sheets":
            selected_sheets = sheet_names[:n_sheets]
            st.session_state.selected_sheets_count = n_sheets
        elif processing_mode == "Last N Sheets":
            selected_sheets = sheet_names[-n_sheets:]
            st.session_state.selected_sheets_count = n_sheets
        elif processing_mode == "Custom Range":
            selected_sheets = sheet_names[start_sheet-1:end_sheet]
            st.session_state.selected_sheets_count = len(selected_sheets)
        
        # Show processing info
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Sheets", total_sheets)
        with col2:
            st.metric("Sheets to Process", st.session_state.selected_sheets_count)
        with col3:
            st.metric("File Size", f"{len(uploaded_file.getvalue()) / 1024:.1f} KB")
        
        # Process button
        if st.button("🚀 Start Extraction", type="primary", use_container_width=True):
            all_results = []
            progress_bar = None
            status_text = None
            
            if show_progress:
                progress_bar = st.progress(0)
                status_text = st.empty()
            
            # Create a container for logs
            log_container = st.container()
            
            with log_container:
                st.markdown("### 📝 Extraction Logs")
                logs_placeholder = st.empty()
                
                logs = []
                start_time = time.time()
                
                for i, sheet_name in enumerate(selected_sheets):
                    if show_progress:
                        progress = (i + 1) / len(selected_sheets)
                        progress_bar.progress(progress)
                        status_text.text(f"Processing sheet {i+1} of {len(selected_sheets)}: {sheet_name}")
                    
                    # Update logs
                    if verbose_mode:
                        logs.append(f"📋 **Processing Sheet {i+1}: {sheet_name}**")
                        logs_placeholder.markdown("\n".join(logs[-5:]))  # Show last 5 logs
                    
                    try:
                        df = read_excel_with_merged_cells(uploaded_file.getvalue(), sheet_name)
                        
                        if not df.empty:
                            sheet_data = extract_invoice_data_from_sheet(df, uploaded_file.name, sheet_name)
                            
                            if sheet_data['items_df'] is not None and not sheet_data['items_df'].empty:
                                all_results.append(sheet_data)
                                
                                if verbose_mode:
                                    logs.append(f"   ✅ Extracted {len(sheet_data['items_df'])} items")
                                    logs.append(f"   👤 Customer: {sheet_data['customer_name']}")
                                    logs_placeholder.markdown("\n".join(logs[-5:]))
                        
                    except Exception as e:
                        if verbose_mode:
                            logs.append(f"   ⚠️ Error: {str(e)[:50]}")
                            logs_placeholder.markdown("\n".join(logs[-5:]))
                        continue
                    
                    # Small delay for UI update
                    time.sleep(0.01)
                
                end_time = time.time()
                processing_time = end_time - start_time
                
                # Combine all results
                if all_results:
                    combined_list = []
                    
                    for result in all_results:
                        items_df = result['items_df'].copy()
                        
                        # Add metadata columns
                        items_df['File Name'] = result['file_name']
                        items_df['Sheet Name'] = result['sheet_name']
                        items_df['Customer Name'] = result['customer_name']
                        items_df['Document Date'] = result['document_date']
                        items_df['Invoice Number'] = result['invoice_number']
                        items_df['Order Number'] = result['order_number']
                        items_df['Subtotal'] = result['subtotal']
                        items_df['VAT Amount'] = result['vat_amount']
                        items_df['Total Amount'] = result['total_amount']
                        
                        combined_list.append(items_df)
                    
                    # Combine all data
                    combined_df = pd.concat(combined_list, ignore_index=True)
                    
                    # Reorder columns
                    column_order = [
                        'File Name', 'Sheet Name', 'Customer Name', 'Document Date',
                        'Invoice Number', 'Order Number', 'Subtotal', 'VAT Amount', 'Total Amount',
                        'No.', 'Description', 'Qty', 'UoM',
                        'Unit Price Excl. VAT', 'VAT %', 'Line Amount Excl. VAT'
                    ]
                    
                    available_cols = [col for col in column_order if col in combined_df.columns]
                    combined_df = combined_df[available_cols]
                    
                    # Store in session state
                    st.session_state.extracted_data = combined_df
                    st.session_state.processing_complete = True
                    
                    # Show success message
                    st.markdown(f'<div class="success-box">✅ Extraction Complete! Processed {len(all_results)} sheets in {processing_time:.1f} seconds</div>', unsafe_allow_html=True)
                    
                    # Display summary metrics
                    st.markdown("### 📊 Extraction Summary")
                    
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric("Sheets Processed", len(all_results))
                    with col2:
                        st.metric("Total Items", len(combined_df))
                    with col3:
                        st.metric("Unique Customers", combined_df['Customer Name'].nunique())
                    with col4:
                        st.metric("Processing Time", f"{processing_time:.1f}s")
                    
                    # Display sample data
                    st.markdown("### 📋 Sample Data")
                    
                    # Create tabs for different views
                    tab1, tab2, tab3 = st.tabs(["Sample Data", "Data Statistics", "Export Options"])
                    
                    with tab1:
                        st.dataframe(combined_df.head(10), use_container_width=True)
                    
                    with tab2:
                        st.markdown("#### Data Overview")
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.markdown("##### Column Information")
                            col_info = pd.DataFrame({
                                'Column': combined_df.columns,
                                'Non-Null Count': combined_df.notna().sum().values,
                                'Data Type': combined_df.dtypes.astype(str).values
                            })
                            st.dataframe(col_info, use_container_width=True)
                        
                        with col2:
                            st.markdown("##### Numerical Summary")
                            if combined_df.select_dtypes(include=[np.number]).shape[1] > 0:
                                st.dataframe(combined_df.describe(), use_container_width=True)
                            else:
                                st.info("No numerical columns found for summary statistics.")
                    
                    with tab3:
                        st.markdown("#### Export Options")
                        
                        # Generate filename with timestamp
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        filename = f"extracted_invoices_{timestamp}.xlsx"
                        
                        # Download button
                        st.markdown("### 📥 Download Extracted Data")
                        st.markdown(get_excel_download_link(combined_df, filename), unsafe_allow_html=True)
                        
                        # Also show CSV option
                        csv = combined_df.to_csv(index=False)
                        b64_csv = base64.b64encode(csv.encode()).decode()
                        csv_filename = f"extracted_invoices_{timestamp}.csv"
                        href_csv = f'<a href="data:file/csv;base64,{b64_csv}" download="{csv_filename}">📥 Download CSV File</a>'
                        st.markdown(href_csv, unsafe_allow_html=True)
                        
                        # Preview of data structure
                        st.markdown("#### Data Structure Preview")
                        st.json({
                            "total_records": len(combined_df),
                            "columns": list(combined_df.columns),
                            "sample_record": combined_df.iloc[0].to_dict() if len(combined_df) > 0 else {}
                        })
                
                else:
                    st.markdown('<div class="warning-box">⚠️ No data was extracted from any sheets. Please check your file format.</div>', unsafe_allow_html=True)
                
                if show_progress:
                    progress_bar.empty()
                    status_text.empty()
    
    except Exception as e:
        st.error(f"Error reading file: {str(e)}")

else:
    # Show welcome message when no file is uploaded
    st.markdown("""
    <div class="info-box">
    <h3>Welcome to the Invoice Data Extractor!</h3>
    <p>This tool helps you extract invoice data from Excel files with the following features:</p>
    <ul>
        <li><strong>Smart Extraction:</strong> Handles merged cells and various formats</li>
        <li><strong>Multi-Sheet Processing:</strong> Process all sheets or selected ranges</li>
        <li><strong>Complete Data:</strong> Extracts customer info, dates, invoice numbers, and all product details</li>
        <li><strong>Easy Export:</strong> Download results as Excel or CSV files</li>
    </ul>
    <p>To get started, upload your Excel file using the uploader above.</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Show example of extracted data format
    st.markdown("### 📋 Expected Output Format")
    
    example_data = pd.DataFrame({
        'File Name': ['example.xlsx', 'example.xlsx'],
        'Sheet Name': ['Sheet1', 'Sheet1'],
        'Customer Name': ['Chandarana Delis-Lavington Branch', 'Chandarana Delis-Lavington Branch'],
        'Document Date': ['4. December 2024', '4. December 2024'],
        'Invoice Number': ['1238685', '1238685'],
        'Order Number': ['287836', '287836'],
        'No.': ['BCH-10212', 'BRC-18502'],
        'Description': ["Brown's Buttery Brie per kg", 'Pecorino per kg'],
        'Qty': [0.32, 0.26],
        'UoM': ['Kilograms', 'Kilograms'],
        'Unit Price Excl. VAT': [2459.98, 3136.00],
        'VAT %': [16.0, 16.0],
        'Line Amount Excl. VAT': [787.19, 815.36]
    })
    
    st.dataframe(example_data, use_container_width=True)

# Footer
st.markdown("---")
st.markdown(
    """
    <div style="text-align: center; color: #666; font-size: 0.9rem;">
    <p>Invoice Data Extractor v1.0 | Built with Streamlit</p>
    <p>© 2024 | Extracts data from Excel invoices with precision</p>
    </div>
    """,
    unsafe_allow_html=True
)
